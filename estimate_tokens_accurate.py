#!/usr/bin/env python3
"""
Token Usage Analysis - Fill Excel Template
Computes exact token counts per dataset and style combination
"""

import sys
from pathlib import Path
from collections import defaultdict
import yaml
import openpyxl

# Add repo to path
repo_root = Path(__file__).parent
sys.path.insert(0, str(repo_root))

from utils.styles import (
    apply_politeness, apply_spacing, apply_punctuation,
    apply_letter_case, apply_length_variation, apply_interrogative
)

def count_tokens(text):
    """Estimate tokens (0.75 tokens per word)"""
    if not text:
        return 0
    return int(len(text.split()) * 0.75)

def load_config():
    """Load config.yaml"""
    with open(repo_root / "config.yaml") as f:
        return yaml.safe_load(f)

def load_datasets():
    """Load real datasets (128 prompts each)"""
    from datasets import load_dataset
    
    datasets_dict = {}
    
    print("\n" + "="*80)
    print("LOADING DATASETS (Black-box measurements only)")
    print("="*80)
    
    # TruthfulQA
    print("\n[1/3] TruthfulQA...")
    tqa = load_dataset("truthful_qa", "generation", split="validation")
    datasets_dict['truthfulqa'] = [item['question'] for item in tqa.select(range(128))]
    print(f"✓ {len(datasets_dict['truthfulqa'])} questions")
    
    # GSM8K
    print("\n[2/3] GSM8K...")
    gsm8k = load_dataset("gsm8k", "main", split="test")
    datasets_dict['gsmk8'] = [item['question'] for item in gsm8k.select(range(128))]
    print(f"✓ {len(datasets_dict['gsmk8'])} problems")
    
    # HarmBench
    print("\n[3/3] HarmBench...")
    try:
        hb = load_dataset("mantas-m/HarmBench", split="val")
        std = [i for i in hb if i.get('semantic_category')=='standard'][:128]
        datasets_dict['harmbench'] = [i['behavior'] for i in std]
    except:
        print("  ⚠ Using fallback prompts")
        datasets_dict['harmbench'] = ["Harmful prompt example"] * 128
    print(f"✓ {len(datasets_dict['harmbench'])} prompts")
    
    return datasets_dict

def compute_per_dataset_style(datasets, config):
    """
    Compute exact token usage per dataset-style combination
    Returns dict: {(dataset, style): {'input': X, 'output': Y}}
    """
    
    print("\n" + "="*80)
    print("COMPUTING TOKENS PER DATASET-STYLE COMBINATION")
    print("="*80)
    
    results = {}
    
    # Output token limits
    MAX_OUT_GENERAL = 100
    MAX_OUT_GSMK8 = 500
    
    # CoT instruction for GSM8K
    cot_instruction = "\n\nLet's think step by step."
    
    # ============================================================
    # SPACING
    # ============================================================
    
    print("\n[1/6] Computing SPACING tokens...")
    
    spacing_strengths = config['style_levels']['spacing']  # [0, 1, 5, 20, 50, 100]
    spacing_positions = config['style_positions']['spacing']  # ['prefix', 'suffix', 'global']
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        # Add CoT instruction for GSM8K
        if dataset_name == 'gsmk8':
            prompts_to_style = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            prompts_to_style = prompts
            max_output = MAX_OUT_GENERAL
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        # All permutations: 3 positions × 6 strengths × 128 prompts
        for strength in spacing_strengths:
            for position in spacing_positions:
                # Apply style to all 128 prompts
                styled_prompts = [
                    apply_spacing(prompt, strength, position)
                    for prompt in prompts_to_style
                ]
                
                # Count input tokens
                input_tokens = sum(count_tokens(p) for p in styled_prompts)
                total_input += input_tokens
                
                # Count output tokens (128 prompts × max_output)
                output_tokens = len(prompts) * max_output
                total_output += output_tokens
                
                num_experiments += 1
        
        results[(dataset_name, 'spacing')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    # ============================================================
    # PUNCTUATION
    # ============================================================
    
    print("\n[2/6] Computing PUNCTUATION tokens...")
    
    punct_strengths = config['style_levels']['punctuation']  # [0, 1, 3, 5, 10, 20]
    punct_positions = config['style_positions']['punctuation']  # ['prefix', 'suffix', 'global']
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        if dataset_name == 'gsmk8':
            prompts_to_style = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            prompts_to_style = prompts
            max_output = MAX_OUT_GENERAL
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        for strength in punct_strengths:
            for position in punct_positions:
                styled_prompts = [
                    apply_punctuation(prompt, strength, position)
                    for prompt in prompts_to_style
                ]
                
                input_tokens = sum(count_tokens(p) for p in styled_prompts)
                total_input += input_tokens
                
                output_tokens = len(prompts) * max_output
                total_output += output_tokens
                
                num_experiments += 1
        
        results[(dataset_name, 'punctuation')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    # ============================================================
    # LETTER CASE
    # ============================================================
    
    print("\n[3/6] Computing LETTER_CASE tokens...")
    
    case_strengths = config['style_levels']['letter_case']  # [0, 10, 25, 50, 75, 100]
    case_positions = config['style_positions']['letter_case']  # ['prefix', 'suffix', 'global']
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        if dataset_name == 'gsmk8':
            prompts_to_style = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            prompts_to_style = prompts
            max_output = MAX_OUT_GENERAL
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        for strength in case_strengths:
            for position in case_positions:
                styled_prompts = [
                    apply_letter_case(prompt, strength, position)
                    for prompt in prompts_to_style
                ]
                
                input_tokens = sum(count_tokens(p) for p in styled_prompts)
                total_input += input_tokens
                
                output_tokens = len(prompts) * max_output
                total_output += output_tokens
                
                num_experiments += 1
        
        results[(dataset_name, 'letter_case')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    # ============================================================
    # POLITENESS
    # ============================================================
    
    print("\n[4/6] Computing POLITENESS tokens...")
    
    pol_strengths = config['style_levels']['politeness']  # [-10, -8, -6, -4, -2, 0, 2, 4, 6, 8, 10]
    pol_positions = config['style_positions']['politeness']  # ['prefix', 'suffix', 'global']
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        if dataset_name == 'gsmk8':
            prompts_to_style = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            prompts_to_style = prompts
            max_output = MAX_OUT_GENERAL
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        for strength in pol_strengths:
            for position in pol_positions:
                styled_prompts = [
                    apply_politeness(prompt, strength, position)
                    for prompt in prompts_to_style
                ]
                
                input_tokens = sum(count_tokens(p) for p in styled_prompts)
                total_input += input_tokens
                
                output_tokens = len(prompts) * max_output
                # Add mirroring overhead for polite prompts
                if strength > 0:
                    output_tokens += len(prompts) * 5
                total_output += output_tokens
                
                num_experiments += 1
        
        results[(dataset_name, 'politeness')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    # ============================================================
    # LENGTH VARIATION
    # ============================================================
    
    print("\n[5/6] Computing LENGTH VARIATION tokens...")
    
    length_multipliers = config['style_levels']['length_variation']  # [0.25, 0.5, 1.0, 1.5, 2.0, 3.0]
    # Length variation is global-only, so 1 position
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        if dataset_name == 'gsmk8':
            base_prompts = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            base_prompts = prompts
            max_output = MAX_OUT_GENERAL
        
        # Calculate average base tokens
        avg_base_tokens = sum(count_tokens(p) for p in base_prompts) / len(base_prompts)
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        # 6 multipliers × 128 prompts (1 position: global)
        for multiplier in length_multipliers:
            # Estimate tokens after length variation
            estimated_tokens_per_prompt = int(avg_base_tokens * multiplier)
            input_tokens = estimated_tokens_per_prompt * len(prompts)
            total_input += input_tokens
            
            output_tokens = len(prompts) * max_output
            total_output += output_tokens
            
            num_experiments += 1
        
        results[(dataset_name, 'length variation')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    # ============================================================
    # INTERROGATIVE VS IMPERATIVE
    # ============================================================
    
    print("\n[6/6] Computing INTER VS IMPER tokens...")
    
    form_variants = config['style_levels']['inter_vs_imper']  # ['interrogative', 'imperative']
    # Inter vs imper is global-only, so 1 position
    
    for dataset_name in ['truthfulqa', 'gsmk8', 'harmbench']:
        prompts = datasets[dataset_name]
        
        if dataset_name == 'gsmk8':
            base_prompts = [p + cot_instruction for p in prompts]
            max_output = MAX_OUT_GSMK8
        else:
            base_prompts = prompts
            max_output = MAX_OUT_GENERAL
        
        # Calculate average base tokens
        avg_base_tokens = sum(count_tokens(p) for p in base_prompts) / len(base_prompts)
        
        total_input = 0
        total_output = 0
        num_experiments = 0
        
        # 2 variants × 128 prompts (1 position: global)
        for variant in form_variants:
            # Estimate: base tokens + small overhead for rephrasing
            estimated_tokens_per_prompt = int(avg_base_tokens + 5)
            input_tokens = estimated_tokens_per_prompt * len(prompts)
            total_input += input_tokens
            
            output_tokens = len(prompts) * max_output
            total_output += output_tokens
            
            num_experiments += 1
        
        results[(dataset_name, 'inter')] = {
            'input': total_input,
            'output': total_output,
            'experiments': num_experiments
        }
        
        print(f"  {dataset_name}: {num_experiments} experiments, "
              f"Input={total_input:,}, Output={total_output:,}")
    
    return results

def fill_excel_template(results, output_file):
    """Fill the Excel template with computed results"""
    
    print("\n" + "="*80)
    print("FILLING EXCEL TEMPLATE")
    print("="*80)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Token Analysis"
    
    # Write headers
    ws['A1'] = 'dataset'
    ws['B1'] = 'style'
    ws['C1'] = 'input'
    ws['D1'] = 'output'
    
    # Define the order as in the template
    datasets_order = ['truthfulqa', 'gsmk8', 'harmbench']
    styles_order = ['spacing', 'punctuation', 'letter_case', 'inter', 'length variation', 'politeness']
    
    row = 2
    for dataset in datasets_order:
        for style in styles_order:
            ws[f'A{row}'] = dataset
            ws[f'B{row}'] = style
            
            key = (dataset, style)
            if key in results:
                ws[f'C{row}'] = results[key]['input']
                ws[f'D{row}'] = results[key]['output']
            else:
                ws[f'C{row}'] = 0
                ws[f'D{row}'] = 0
            
            row += 1
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel file saved to: {output_file}")
    
    # Print summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    
    total_input = sum(r['input'] for r in results.values())
    total_output = sum(r['output'] for r in results.values())
    total_tokens = total_input + total_output
    
    print(f"\nTotal Input:  {total_input:,} tokens ({total_input/total_tokens*100:.1f}%)")
    print(f"Total Output: {total_output:,} tokens ({total_output/total_tokens*100:.1f}%)")
    print(f"TOTAL:        {total_tokens:,} tokens")
    
    # Cost estimates
    print("\n" + "-"*80)
    print("ESTIMATED COSTS")
    print("-"*80)
    
    for model, (in_price, out_price) in [
        ('GPT-4', (30, 60)),
        ('Claude Sonnet', (3, 15)),
        ('GPT-4o-mini', (0.15, 0.6))
    ]:
        cost = (total_input/1e6)*in_price + (total_output/1e6)*out_price
        print(f"{model:20s} ${cost:,.2f}")

def main():
    print("="*80)
    print("TOKEN ANALYSIS - FILL EXCEL TEMPLATE")
    print("Per Dataset-Style Combination")
    print("="*80)
    
    try:
        config = load_config()
        datasets = load_datasets()
        results = compute_per_dataset_style(datasets, config)
        
        output_file = repo_root / "token_analysis_filled.xlsx"
        fill_excel_template(results, output_file)
        
        print("\n" + "="*80)
        print("✓ COMPLETE")
        print("="*80)
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
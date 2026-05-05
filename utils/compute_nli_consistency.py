"""
utils/compute_nli_consistency.py
---------------------------------
Computes NLI-based factual consistency metrics for all runs in the checklist
and appends them to each full_results_all_models.csv in-place.

New columns added
-----------------
  nli_contradiction        : P(contradiction | response_orig, response_pert) in [0, 1]
  nli_entailment           : P(entailment    | response_orig, response_pert) in [0, 1]
  delta_nli_contradiction  : signed delta relative to within-group baseline
  delta_nli_entailment     : signed delta relative to within-group baseline

Metric justification
--------------------
NLI contradiction between a model's baseline response and its perturbed-prompt
response captures factual instability that BLEU and BERTScore cannot detect:
two responses can be surface-dissimilar but logically consistent, or
surface-similar but factually contradictory. This provides a logit-free proxy
for confidence stability applicable to all models, including closed-source ones
that do not expose log-probabilities.

Model  : cross-encoder/nli-deberta-v3-large  (Laurer et al., 2022, EMNLP)
Cited in: Kuhn et al. 2023 (ICLR), Farquhar et al. 2024 (Nature),
          Welleck et al. 2019 (ACL), Lin et al. 2022 (ACL)

Delta logic  (mirrors compute_delta_metrics.py)
-----------------------------------------------
For each (model, prompt_id, place) group, the baseline row is identified by
the style-specific baseline strength:
  Politeness           : strength == 0
  Spacing / Punct. / Casing : strength == 0
  Length               : strength == 1.0
  Form (inter_vs_imper): strength == 'original'  (fallback: 'interrogative')

delta_nli_* = metric_at_variant - metric_at_baseline

Usage
-----
Run from project root:
    python utils/compute_nli_consistency.py --mode open
    python utils/compute_nli_consistency.py --mode closed

    --mode open   : reads experiment_checklist_filled.xlsx
    --mode closed : reads experiment_checklist_closed_filled.xlsx
    --force       : recompute even if columns already present
    --batch-size  : NLI inference batch size (default 32)
    --device      : cuda device index (default 0)
"""

import argparse
import sys
import warnings
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import torch
import openpyxl
from tqdm import tqdm
from transformers import AutoTokenizer, AutoModelForSequenceClassification

warnings.filterwarnings("ignore")

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--mode",       choices=["open", "closed"], required=True,
                    help="open = open-source checklist, closed = closed-source checklist")
parser.add_argument("--force",      action="store_true",
                    help="recompute even if NLI columns already present")
parser.add_argument("--batch-size", type=int, default=32)
parser.add_argument("--device",     type=int, default=0)
args = parser.parse_args()

CHECKLIST_MAP = {
    "open":   Path("experiment_checklist_filled.xlsx"),
    "closed": Path("experiment_checklist_closed_filled.xlsx"),
}
CHECKLIST_PATH = CHECKLIST_MAP[args.mode]
NLI_COLS = ["nli_contradiction", "nli_entailment",
            "delta_nli_contradiction", "delta_nli_entailment"]

# ── Baseline strength per style ───────────────────────────────────────────────
# These must match compute_delta_metrics.py
BASELINE_STRENGTH = {
    "Polite.":  0,        # politeness
    "Spacing":  0,        # surface noise
    "Punct.":   0,        # surface noise
    "Casing":   0,        # surface noise
    "Length":   1.0,      # length_variation multiplier = 1.0 means no change
    "Form":     None,     # special: 'original' then fallback 'interrogative'
}

ALL_STYLES = set(BASELINE_STRENGTH.keys())


def get_baseline_mask(df: pd.DataFrame, style: str) -> pd.Series:
    """Return a boolean mask marking the baseline rows for this style."""
    if style == "Form":
        strength_col = df["strength"].astype(str).str.strip().str.lower()
        if (strength_col == "original").any():
            return strength_col == "original"
        return strength_col == "interrogative"
    else:
        target = BASELINE_STRENGTH[style]
        try:
            return df["strength"].astype(float) == float(target)
        except (ValueError, TypeError):
            return df["strength"].astype(str) == str(target)


# ── Read checklist ─────────────────────────────────────────────────────────────
if not CHECKLIST_PATH.exists():
    sys.exit(f"[ERROR] Checklist not found: {CHECKLIST_PATH}\n"
             f"        Run fill_checklist.py / fill_checklist_closed.py first.")

print(f"\n[INFO] Reading checklist: {CHECKLIST_PATH}")
wb = openpyxl.load_workbook(CHECKLIST_PATH, read_only=True, data_only=True)
ws = wb.active

entries = []   # list of {path, style, model, dataset}
cur_model = cur_dataset = None

for r in range(3, ws.max_row + 1):
    v2 = ws.cell(r, 2).value
    v3 = ws.cell(r, 3).value
    v4 = ws.cell(r, 4).value
    v5 = ws.cell(r, 5).value

    if v2 is not None: cur_model   = str(v2).strip()
    if v3 is not None:
        v3s = str(v3).strip()
        if v3s.startswith("▶"):
            continue
        cur_dataset = v3s

    if v4 is None or str(v4).strip() not in ALL_STYLES:
        continue
    if cur_model is None or cur_dataset is None:
        continue

    style = str(v4).strip()
    path  = str(v5).strip() if v5 is not None else ""

    if not path or path.upper() == "NOT FOUND":
        continue

    entries.append({
        "path":    Path(path),
        "style":   style,
        "model":   cur_model,
        "dataset": cur_dataset,
    })

wb.close()

# Deduplicate paths (same run folder may appear multiple times for
# different models in open-source checklists)
unique_paths = {}   # path -> style  (style is the same for all rows of a run)
for e in entries:
    if e["path"] not in unique_paths:
        unique_paths[e["path"]] = e["style"]

print(f"[INFO] {len(entries)} checklist rows → {len(unique_paths)} unique run folders")

# Filter out already-computed (unless --force)
if not args.force:
    to_process = {}
    skipped = 0
    for path, style in unique_paths.items():
        csv_path = path / "full_results_all_models.csv"
        if not csv_path.exists():
            print(f"  [WARN] CSV not found: {csv_path}")
            continue
        try:
            header = pd.read_csv(csv_path, nrows=0).columns.tolist()
        except Exception:
            to_process[path] = style
            continue
        if all(c in header for c in NLI_COLS):
            skipped += 1
        else:
            to_process[path] = style
    if skipped:
        print(f"[INFO] Skipping {skipped} folders (already computed). Use --force to recompute.")
else:
    to_process = {p: s for p, s in unique_paths.items()
                  if (p / "full_results_all_models.csv").exists()}

if not to_process:
    print("[INFO] Nothing to compute.")
    sys.exit(0)

print(f"[INFO] Will process {len(to_process)} run folders.\n")

# ── Load NLI model ─────────────────────────────────────────────────────────────
MODEL_NAME = "cross-encoder/nli-deberta-v3-large"
print(f"[INFO] Loading NLI model: {MODEL_NAME}")

device = torch.device(f"cuda:{args.device}" if torch.cuda.is_available() else "cpu")
print(f"[INFO] Using device: {device}")

tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
model = AutoModelForSequenceClassification.from_pretrained(MODEL_NAME)
model.eval().to(device)

# DeBERTa NLI label mapping: check which index maps to 'contradiction'
# Standard for cross-encoder/nli-deberta-v3-large: {0: contradiction, 1: entailment, 2: neutral}
# but verify from model config
id2label = model.config.id2label
print(f"[INFO] Model label mapping: {id2label}")

# Find label indices
label_lower = {v.lower(): k for k, v in id2label.items()}
CONTRA_IDX  = label_lower.get("contradiction", 0)
ENTAIL_IDX  = label_lower.get("entailment",    1)
print(f"[INFO] contradiction idx={CONTRA_IDX}, entailment idx={ENTAIL_IDX}\n")


def batch_nli_scores(premises: list[str], hypotheses: list[str],
                     batch_size: int) -> tuple[np.ndarray, np.ndarray]:
    """
    Returns arrays of shape (N,) for contradiction and entailment probabilities.
    Input: premise = response_orig, hypothesis = response_pert
    """
    all_contra = []
    all_entail = []

    for start in range(0, len(premises), batch_size):
        batch_p = premises[start : start + batch_size]
        batch_h = hypotheses[start : start + batch_size]

        enc = tokenizer(
            batch_p, batch_h,
            padding=True, truncation=True,
            max_length=512, return_tensors="pt"
        ).to(device)

        with torch.no_grad():
            logits = model(**enc).logits          # (B, num_labels)
            probs  = torch.softmax(logits, dim=-1).cpu().numpy()

        all_contra.extend(probs[:, CONTRA_IDX].tolist())
        all_entail.extend(probs[:, ENTAIL_IDX].tolist())

    return np.array(all_contra), np.array(all_entail)


# ── Main loop ─────────────────────────────────────────────────────────────────
outer_bar = tqdm(
    to_process.items(),
    total=len(to_process),
    desc=f"[{args.mode.upper()}] Run folders",
    unit="folder",
    ncols=100,
    colour="green",
)

n_updated = 0
n_errors  = 0

for run_path, style in outer_bar:
    csv_path = run_path / "full_results_all_models.csv"
    outer_bar.set_postfix_str(f"{run_path.name[:40]}")

    # ── Load ──────────────────────────────────────────────────────────────────
    try:
        df = pd.read_csv(csv_path, low_memory=False)
    except Exception as exc:
        tqdm.write(f"  [ERR] Cannot read {csv_path}: {exc}")
        n_errors += 1
        continue

    if df.empty:
        tqdm.write(f"  [SKIP] Empty CSV: {csv_path}")
        continue

    # Ensure response columns are strings, fill NaN with empty string
    df["response_orig"] = df["response_orig"].fillna("").astype(str)
    df["response_pert"] = df["response_pert"].fillna("").astype(str)

    premises    = df["response_orig"].tolist()
    hypotheses  = df["response_pert"].tolist()

    # ── NLI scores ────────────────────────────────────────────────────────────
    inner_bar = tqdm(
        total=len(df),
        desc=f"  NLI  {run_path.name[:30]}",
        unit="row",
        leave=False,
        ncols=100,
    )

    contra_arr = np.zeros(len(df))
    entail_arr = np.zeros(len(df))

    for start in range(0, len(df), args.batch_size):
        end     = min(start + args.batch_size, len(df))
        batch_p = premises[start:end]
        batch_h = hypotheses[start:end]

        enc = tokenizer(
            batch_p, batch_h,
            padding=True, truncation=True,
            max_length=512, return_tensors="pt"
        ).to(device)

        with torch.no_grad():
            logits = model(**enc).logits
            probs  = torch.softmax(logits, dim=-1).cpu().numpy()

        contra_arr[start:end] = probs[:, CONTRA_IDX]
        entail_arr[start:end] = probs[:, ENTAIL_IDX]
        inner_bar.update(end - start)

    inner_bar.close()

    df["nli_contradiction"] = contra_arr
    df["nli_entailment"]    = entail_arr

    # ── Delta computation ─────────────────────────────────────────────────────
    # For each (model, prompt_id, place) group, subtract the baseline-strength
    # row's NLI score from every row in the group.
    baseline_mask = get_baseline_mask(df, style)

    df["delta_nli_contradiction"] = np.nan
    df["delta_nli_entailment"]    = np.nan

    group_cols = ["model", "prompt_id", "place"]
    # Only use columns that actually exist in the dataframe
    group_cols = [c for c in group_cols if c in df.columns]

    if not group_cols:
        # Fallback: no grouping — treat every row's own baseline_mask
        tqdm.write(f"  [WARN] No group columns found in {csv_path.name}, using global baseline")
        base_contra = df.loc[baseline_mask, "nli_contradiction"].mean()
        base_entail = df.loc[baseline_mask, "nli_entailment"].mean()
        df["delta_nli_contradiction"] = df["nli_contradiction"] - base_contra
        df["delta_nli_entailment"]    = df["nli_entailment"]    - base_entail
    else:
        for group_key, group_idx in df.groupby(group_cols).groups.items():
            grp = df.loc[group_idx]
            base_rows = grp[get_baseline_mask(grp, style)]

            if base_rows.empty:
                # Baseline not found in this group — set delta to NaN
                tqdm.write(
                    f"  [WARN] No baseline row found for group "
                    f"{dict(zip(group_cols, [group_key] if not isinstance(group_key, tuple) else group_key))} "
                    f"in {run_path.name}"
                )
                continue

            base_contra = float(base_rows["nli_contradiction"].iloc[0])
            base_entail = float(base_rows["nli_entailment"].iloc[0])

            df.loc[group_idx, "delta_nli_contradiction"] = (
                df.loc[group_idx, "nli_contradiction"] - base_contra
            )
            df.loc[group_idx, "delta_nli_entailment"] = (
                df.loc[group_idx, "nli_entailment"] - base_entail
            )

    # ── Save back ─────────────────────────────────────────────────────────────
    try:
        df.to_csv(csv_path, index=False)
        n_updated += 1
        tqdm.write(
            f"  [OK]  {run_path.name} | rows={len(df)} | "
            f"mean_contra={contra_arr.mean():.4f} | "
            f"style={style}"
        )
    except Exception as exc:
        tqdm.write(f"  [ERR] Cannot write {csv_path}: {exc}")
        n_errors += 1

# ── Summary ───────────────────────────────────────────────────────────────────
print(f"\n{'='*64}")
print(f"  Mode    : {args.mode}")
print(f"  Updated : {n_updated} run folders")
print(f"  Errors  : {n_errors}")
print(f"  Columns added: {NLI_COLS}")
print(f"  Model   : {MODEL_NAME}")
print(f"{'='*64}\n")

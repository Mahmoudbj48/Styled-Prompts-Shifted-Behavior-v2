"""
fill_checklist_closed.py
------------------------
Fills experiment_checklist_to_fill_closed_models.xlsx for closed-source models.

Run from project root:
    python fill_checklist_closed.py

Scans: results/closed_models/run_<model>_<dataset>_<style>_<timestamp>/full_results_all_models.csv
Reads:  experiment_checklist_to_fill_closed_models.xlsx
Writes: experiment_checklist_closed_filled.xlsx

For each (model x dataset x style) row in the checklist:
- Finds the matching run folder (most rows wins if duplicates exist)
- Fills col E (path) and col F (# prompts) with green formatting
- Marks missing combinations in orange
Does NOT move or delete any folders.
"""

import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from collections import defaultdict

CLOSED_DIR    = Path("results/closed_models")
CHECKLIST_IN  = Path("experiment_checklist_to_fill_closed_models.xlsx")
CHECKLIST_OUT = Path("experiment_checklist_closed_filled.xlsx")

STYLE_MAP = {
    "inter_vs_imper":   "Form",
    "length_variation": "Length",
    "letter_case":      "Casing",
    "politeness":       "Polite.",
    "punctuation":      "Punct.",
    "spacing":          "Spacing",
}

DATASET_MAP = {
    "truthful_qa":       "TruthfulQA",
    "natural_questions": "Natural Questions",
    "alpaca":            "Alpaca",
    "simpleqa_verified": "SimpleQA Verified",
    "trivia_qa":         "TriviaQA",
    "hotpot_qa":         "HotpotQA",
}

MODEL_ALIASES = {
    "GPT-5.4":           ["gpt-5.4", "gpt5.4", "gpt-5_4"],
    "Gemini-2.5-Flash":  ["gemini-2.5-flash", "gemini-flash-2.5"],
    "Claude-Sonnet-4.6": ["claude-sonnet-4-6", "claude-sonnet-4.6", "claude-sonnet-46"],
}

_MODEL_LOOKUP = {}
for label, aliases in MODEL_ALIASES.items():
    _MODEL_LOOKUP[label.lower()] = label
    for a in aliases:
        _MODEL_LOOKUP[a.lower()] = label

def normalise_model(raw):
    return _MODEL_LOOKUP.get(raw.strip().lower(), raw.strip())

ALL_STYLES = set(STYLE_MAP.values())

# ── Scan results/closed_models/ ───────────────────────────────────────────────
# candidates[(style, model, dataset)] = list of {run_dir, n_rows, n_prompts}
candidates = defaultdict(list)

if not CLOSED_DIR.is_dir():
    sys.exit(f"[ERROR] {CLOSED_DIR} not found.")

for run_dir in sorted(CLOSED_DIR.iterdir()):
    if not run_dir.is_dir() or run_dir.name in {"trash", "empty_results"}:
        continue
    csv_path = run_dir / "full_results_all_models.csv"
    if not csv_path.exists():
        continue

    # Detect style (longest match first)
    style_label = None
    for slug in sorted(STYLE_MAP, key=len, reverse=True):
        if f"_{slug}_" in run_dir.name:
            style_label = STYLE_MAP[slug]
            break
    if style_label is None:
        print(f"  [SKIP] Cannot detect style from {run_dir.name}")
        continue

    # Detect dataset (longest match first)
    dataset_label = None
    for slug in sorted(DATASET_MAP, key=len, reverse=True):
        if f"_{slug}_" in run_dir.name:
            dataset_label = DATASET_MAP[slug]
            break
    if dataset_label is None:
        print(f"  [SKIP] Cannot detect dataset from {run_dir.name}")
        continue

    try:
        df = pd.read_csv(csv_path, usecols=["model", "prompt_id"], low_memory=False)
    except Exception as exc:
        print(f"  [ERR] {run_dir.name}: {exc}")
        continue

    for model_raw, grp in df.groupby("model"):
        key = (style_label, normalise_model(model_raw), dataset_label)
        candidates[key].append({
            "run_dir":   run_dir,
            "n_rows":    len(grp),
            "n_prompts": grp["prompt_id"].nunique(),
        })

# Pick best (most rows, tie → newest folder name) per key
best = {}
for key, cands in candidates.items():
    best[key] = sorted(cands, key=lambda c: (c["n_rows"], c["run_dir"].name), reverse=True)[0]

# ── Fill checklist ────────────────────────────────────────────────────────────
if not CHECKLIST_IN.exists():
    sys.exit(f"[ERROR] {CHECKLIST_IN} not found.")

wb = openpyxl.load_workbook(CHECKLIST_IN)
ws = wb.active

DONE_BG  = PatternFill("solid", fgColor="C6EFCE")
DONE_FT  = Font(name="Arial", size=9, color="276221")
DONE_NUM = Font(name="Arial", size=9, bold=True, color="276221")
MISS_BG  = PatternFill("solid", fgColor="FCE4D6")
MISS_FT  = Font(name="Arial", size=9, color="C55A11", italic=True)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=False)

cur_model = cur_dataset = None
filled = missing = 0

for row in range(3, ws.max_row + 1):
    v2 = ws.cell(row, 2).value
    v3 = ws.cell(row, 3).value
    v4 = ws.cell(row, 4).value

    if v2 is not None: cur_model   = str(v2).strip()
    if v3 is not None:
        v3s = str(v3).strip()
        if v3s.startswith("▶"):
            continue
        cur_dataset = v3s

    if v4 is None or str(v4).strip() not in ALL_STYLES:
        continue
    if cur_model is None or cur_dataset is None:
        continue

    key  = (str(v4).strip(), cur_model, cur_dataset)
    info = best.get(key)
    pc   = ws.cell(row, 5)
    nc   = ws.cell(row, 6)

    if info:
        pc.value = info["run_dir"].as_posix()
        pc.fill  = DONE_BG; pc.font = DONE_FT; pc.alignment = LEFT
        nc.value = info["n_prompts"]
        nc.fill  = DONE_BG; nc.font = DONE_NUM; nc.alignment = CENTER
        filled += 1
    else:
        pc.value = "NOT FOUND"
        pc.fill  = MISS_BG; pc.font = MISS_FT; pc.alignment = CENTER
        nc.value = "—"
        nc.fill  = MISS_BG; nc.font = MISS_FT; nc.alignment = CENTER
        missing += 1

wb.save(CHECKLIST_OUT)
print(f"[OK]  {CHECKLIST_OUT}")
print(f"      Filled:    {filled}")
print(f"      Not found: {missing}")